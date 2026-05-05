import subprocess
import sys
import os
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
import requests
import cv2
import base64
import json
import re
import numpy as np
from openpyxl import Workbook
from datetime import datetime


# =========================
# CONFIG
# =========================
OLLAMA_URL = "http://192.168.1.83:11434/v1/chat/completions"
MODEL_NAME = "qwen3-vl:4b"
IMAGE_FOLDER = "output"
MAX_WORKERS = 2
SYSTEM_PROMPT = """Bạn là AI chuyên trích xuất thông tin từ văn bản hành chính Việt Nam.

Nhiệm vụ: đọc nội dung được cung cấp và trả về DUY NHẤT 1 JSON object hợp lệ.

TUYỆT ĐỐI:
- Không thêm giải thích.
- Không thêm markdown.
- Không thêm ký tự ngoài JSON.
- Không suy đoán.
- Không tự bịa.

Schema bắt buộc:

{  
  "loai_vb": string,
  "domat": string|null,
  "so_vb": string|null,
  "ngay_banhanh": string|null,
  "trichyeu": string|null
}

ĐỊNH NGHĨA RÕ TỪNG TRƯỜNG:

1) so_vb (SỐ VĂN BẢN):
- Là số và ký hiệu chính thức của văn bản.
- Thường có dạng như: 123/QĐ-UBND, 45/2024/CV-ABC, 789/TB-SGDĐT.
- Thường nằm ở phần đầu văn bản, gần tiêu đề.
- KHÔNG phải số trang.
- KHÔNG phải số phụ lục.
- KHÔNG phải số thứ tự dòng.
- Không tự thêm năm nếu văn bản không ghi.
- Nếu không tìm thấy rõ ràng → null.

2) ngay_banhanh:
- Là ngày ban hành văn bản.
- Chuẩn hóa về định dạng dd/mm/yyyy.
- Nếu thiếu ngày hoặc tháng hoặc năm → null.
- Không tự suy đoán năm.

3) loai_vb:
Là phân loại văn bản
Chỉ được chọn MỘT trong các giá trị sau: Kế hoạch, Quyết định, Công văn, Báo cáo, Báo cáo đề xuất, Tờ trình, Phương án, Phân công lực lượng, Quy định, Thông báo, Điện, Thư cảm ơn, Hướng dẫn, Phiếu xử lý VB, Phiếu xin ý kiến, Chỉ thị, Giấy mời, Nghị quyết, Chương trình
Nếu không thuộc danh sách trên → "Khác".
Không được tự tạo loại mới.

4) domat:
Là độ mật của văn bản, thường được đóng dấu đỏ hoặc dấu đen hình chữ nhật ở phía bên trái.
Chỉ trả về một trong các giá trị:
- Tuyệt mật
- Tối mật
- Mật
Nếu văn bản không ghi rõ độ mật → null.
Không suy diễn.

5) trichyeu:
- Là nội dung ngắn gọn trong 1 câu thể hiện tên văn bản.
- Chỉ trích đúng nội dung xuất hiện trong văn bản. KHÔNG tự tóm tắt nội dung để tạo trích yếu.
- Thường có vị trí sau tên loại văn bản (đối với văn bản được phân loại cụ thể bên trên) hoặc có vị trí dưới số văn bản (đối với văn bản thuộc loại Công văn, thường bắt đầu bằng chữ "V/v" hoặc "Về việc").
- Nếu không tìm thấy trích yếu rõ ràng → null.
- Không viết lại.
- Không tóm tắt.
- Nếu không xác định rõ → null.

QUY TẮC BẮT BUỘC:
- Nếu không chắc chắn về bất kỳ trường nào → trả về null cho trường đó.
- Chỉ dựa trên nội dung được cung cấp.
- Kết quả phải là JSON hợp lệ."""

result = []

# Tạo session dùng chung để reuse connection
session = requests.Session()


# =========================
# 1️⃣ XỬ LÝ ẢNH
# =========================
def img_process(image_path):
    
    img = cv2.imread(image_path)
    if img is None:
        return None
    
    h, w = img.shape[:2]
    #print("Kích thước:", w, "x", h)
    cropped = img[0:int(h * 0.5), :]

    ch, cw = cropped.shape[:2]

    max_width = 900
    if cw > max_width:
        ratio = max_width / cw
        cropped = cv2.resize(cropped, (max_width, int(ch * ratio)))

    _, buffer = cv2.imencode(
        ".jpg",
        cropped,
        [int(cv2.IMWRITE_JPEG_QUALITY), 95]
    )

    return base64.b64encode(buffer).decode("utf-8")


# =========================
# 2️⃣ OCR
# =========================
def process_single_image(data):

    filename, img_base64 = data

    payload = {
        "model": MODEL_NAME,
        "messages": [
            {
                "role": "system",
                "content": SYSTEM_PROMPT
            },
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": "Hãy trích xuất thông tin văn bản và trả về JSON."
                    },
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/jpeg;base64,{img_base64}"
                        }
                    }
                ]
            }
        ],
        "stream": False,
        "max_tokens": 500,
        "keep_alive": "5m",
        "top_p": 0.1
    }

    try:
        response = session.post(OLLAMA_URL, json=payload, timeout=60)
        response.raise_for_status()

        raw_text = response.json()["choices"][0]["message"]["content"]

        # Tối ưu parse JSON nhanh hơn regex
        start = raw_text.find("{")
        end = raw_text.rfind("}") + 1

        if start == -1 or end == -1:
            print(f"Không tìm thấy JSON: {filename}")
            return

        data_json = json.loads(raw_text[start:end])
        data_json["filename"] = filename

        result.append(data_json)

        print(f"OCR xong: {filename}")

    except Exception as e:
        print(f"Lỗi OCR {filename}: {e}")


# =========================
# 3️⃣ CHUẨN HÓA
# =========================
def result_process():

    for item in result:

        # Chuẩn hóa null
        for key, value in item.items():
            if value in ["null", "", "None"]:
                item[key] = None

        # Chuẩn hóa ngày ban hành
        if item.get("ngay_banhanh"):
            item["ngay_banhanh"] = item["ngay_banhanh"].strip()
            if not re.match(r"\d{2}/\d{2}/\d{4}$", item["ngay_banhanh"]):
                item["ngay_banhanh"] = None

        # Chuẩn hóa số văn bản
        if item.get("so_vb"):
            item["so_vb"] = item["so_vb"].strip()

        # ✅ Chuẩn hóa độ mật
        domat = item.get("domat")

        if not domat:
            item["domat"] = "Không"
        else:
            domat_clean = domat.strip().lower()

            if "tuyệt" in domat_clean:
                item["domat"] = "A"
            elif "tối" in domat_clean:
                item["domat"] = "B"
            elif "mật" in domat_clean:
                item["domat"] = "C"
            else:
                item["domat"] = "Không"

def export_to_excel(data, output_file="ket_qua_ocr.xlsx"):

    wb = Workbook()
    ws = wb.active
    ws.title = "DanhSachVB"

    # Header
    ws.append(["loai_vb", "domat", "so_vb", "ngay_banhanh", "trichyeu"])

    for item in data:

        loai_vb = item.get("loai_vb") or ""
        domat = item.get("domat") or ""
        so_vb = item.get("so_vb") or ""
        trichyeu = item.get("trichyeu") or ""

        # Xử lý ngày
        ngay_dt = None
        ngay_str = item.get("ngay_banhanh")
        #print("RAW:", repr(ngay_str))

        if ngay_str:
            try:
                ngay_dt = datetime.strptime(ngay_str.strip(), "%d/%m/%Y")
            except Exception as e:
                print("Lỗi convert ngày:", ngay_str)

        ws.append([loai_vb, domat, so_vb, ngay_dt, trichyeu])

        # Nếu có ngày hợp lệ thì format
        if ngay_dt:
            cell = ws.cell(row=ws.max_row, column=4)
            cell.number_format = "DD/MM/YYYY"

    wb.save(output_file)
    print("Đã xuất:", output_file)


# =========================
# 4️⃣ MAIN
# =========================
def main():

    start_time = time.time()

    if not os.path.exists(IMAGE_FOLDER):
        print("Thư mục ảnh không tồn tại.")
        return

    image_files = [
        f for f in os.listdir(IMAGE_FOLDER)
        if f.lower().endswith((".jpg", ".jpeg", ".png"))
    ]

    if not image_files:
        print("Không có ảnh trong thư mục.")
        return

    print(f"Tìm thấy {len(image_files)} ảnh.")
    
    # BƯỚC 1: Encode ảnh trước (CPU)
    encoded_data = []

    for filename in image_files:
        path = os.path.join(IMAGE_FOLDER, filename)
        img_base64 = img_process(path)
        if img_base64:
            encoded_data.append((filename, img_base64))

    # BƯỚC 2: Gửi API song song
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = [executor.submit(process_single_image, d) for d in encoded_data]

        for future in as_completed(futures):
            future.result()

    result_process()

    #print(json.dumps(result, indent=2, ensure_ascii=False))
    #print(type(result[0]["ngay_banhanh"]))
    print(f"\n⏳ Tổng thời gian: {time.time() - start_time:.2f} giây")


if __name__ == "__main__":
    main()
    export_to_excel(result)