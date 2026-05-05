import subprocess
import sys
import os
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
import requests
import base64
import json
import re
from openpyxl import Workbook
from datetime import datetime
from PIL import Image
import io

# Lấy thư mục gốc (ngoài thư mục data)[cite: 1]
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
IMAGE_FOLDER = os.path.join(BASE_DIR, "output")
# =========================
# CONFIG
# =========================
OLLAMA_URL = "http://ollama.tbasse.freeddns.org/v1/chat/completions"
MODEL_NAME = "ocr"
MAX_WORKERS = 2
SYSTEM_PROMPT = """Bạn là AI chuyên trích xuất thông tin từ văn bản hành chính Việt Nam.

Nhiệm vụ: đọc nội dung được cung cấp và trả về DUY NHẤT 1 JSON object hợp lệ.

TUYỆT ĐỐI:
- Không thêm giải thích.
- Không thêm markdown.
- Không thêm ký tự ngoài JSON.
- Không suy đoán.
- Không tự bịa.

Schema bắt buộc:

{  
  "loai_vb": string,
  "domat": string|null,
  "so_vb": string|null,
  "ngay_banhanh": string|null,
  "trichyeu": string|null
}

ĐỊNH NGHĨA RÕ TỪNG TRƯỜNG:

1) so_vb (SỐ VĂN BẢN):
- Là số và ký hiệu chính thức của văn bản.
- Thường có dạng như: 123/QĐ-UBND, 45/2024/CV-ABC, 789/TB-SGDĐT.
- Thường nằm ở phần đầu văn bản, gần tiêu đề.
- KHÔNG phải số trang.
- KHÔNG phải số phụ lục.
- KHÔNG phải số thứ tự dòng.
- Không tự thêm năm nếu văn bản không ghi.
- Nếu không tìm thấy rõ ràng → null.

2) ngay_banhanh:
- Là ngày ban hành văn bản.
- Chuẩn hóa về định dạng dd/mm/yyyy.
- Nếu thiếu ngày hoặc tháng hoặc năm → null.
- Không tự suy đoán năm.

3) loai_vb:
Là phân loại văn bản
Chỉ được chọn MỘT trong các giá trị sau: Kế hoạch, Quyết định, Công văn, Báo cáo, Báo cáo đề xuất, Tờ trình, Phương án, Phân công lực lượng, Quy định, Thông báo, Điện, Thư cảm ơn, Hướng dẫn, Phiếu xử lý VB, Phiếu xin ý kiến, Chỉ thị, Giấy mời, Nghị quyết, Chương trình
Nếu không thuộc danh sách trên → "Khác".
Không được tự tạo loại mới.

4) domat:
Là độ mật của văn bản, thường được đóng dấu đỏ hoặc dấu đen hình chữ nhật ở phía bên trái.
Chỉ trả về một trong các giá trị:
- Tuyệt mật
- Tối mật
- Mật
Nếu văn bản không ghi rõ độ mật → null.
Không suy diễn.

5) trichyeu:
- Là nội dung ngắn gọn trong 1 câu thể hiện tên văn bản.
- Chỉ trích đúng nội dung xuất hiện trong văn bản. KHÔNG tự tóm tắt nội dung để tạo trích yếu.
- Thường có vị trí sau tên loại văn bản (đối với văn bản được phân loại cụ thể bên trên) hoặc có vị trí dưới số văn bản (đối với văn bản thuộc loại Công văn, thường bắt đầu bằng chữ "V/v" hoặc "Về việc").
- Nếu không tìm thấy trích yếu rõ ràng → null.
- Không viết lại.
- Không tóm tắt.
- Nếu không xác định rõ → null.
- Nếu trong nội dung trích yếu có dấu ngoặc kép, hãy loại bỏ tất cả dấu ngoặc trong phần giá trị của trích yếu để đảm bảo JSON hợp lệ.

QUY TẮC BẮT BUỘC:
- Nếu không chắc chắn về bất kỳ trường nào → trả về null cho trường đó.
- Chỉ dựa trên nội dung được cung cấp.
- Kết quả phải là JSON hợp lệ."""

result = []
session = requests.Session()

# =========================
# 1️⃣ XỬ LÝ ẢNH
# =========================
def img_process(image_path):
    try:
        with Image.open(image_path) as img:
            w, h = img.size
            cropped = img.crop((0, 0, w, int(h * 0.5)))
            max_width = 900
            if cropped.width > max_width:
                ratio = max_width / cropped.width
                new_height = int(cropped.height * ratio)
                cropped = cropped.resize((max_width, new_height))
            buffer = io.BytesIO()
            cropped.save(buffer, format="JPEG", quality=95)
            return base64.b64encode(buffer.getvalue()).decode("utf-8")
    except:
        return None

# =========================
# 2️⃣ LOG & CLEAN JSON
# =========================
def log_console(filename, message, type="INFO"):
    """Hàm log tập trung để sau này bạn dễ dàng redirect vào GUI Console"""
    timestamp = datetime.now().strftime("%H:%M:%S")
    print(f"[{timestamp}] [{type}] [{filename}] {message}")

def fix_json_quotes(raw_text):
    # Trích xuất phần JSON
    match = re.search(r'\{.*\}', raw_text, re.DOTALL)
    if not match: return None
    text = match.group(0)
    
    # Bước 1: Thử parse trực tiếp (nếu AI ngoan, không lỗi)
    try:
        return text
    except: pass

    # Bước 2: Nếu lỗi, dùng regex sửa các lỗi ngoặc kép lồng phổ biến
    # Tìm các đoạn nằm giữa hai dấu phẩy hoặc ngoặc nhọn mà có dấu ngoặc kép thừa
    # Cách đơn giản nhất để xóa dấu ngoặc kép nội dung:
    def clean_illegal_quotes(match):
        full_pair = match.group(0)
        # Giữ lại ngoặc kép ở đầu và cuối giá trị, xóa hết bên trong
        parts = full_pair.split(':')
        key = parts[0]
        val = ':'.join(parts[1:])
        
        start_q = val.find('"')
        end_q = val.rfind('"')
        if start_q != -1 and end_q != -1:
            content = val[start_q+1:end_q].replace('"', '')
            return f'{key}: "{content}"'
        return full_pair

    # Sửa lỗi trên toàn bộ text trước khi parse
    fixed_text = re.sub(r'("[^"]*"\s*:\s*".*?")', clean_illegal_quotes, text, flags=re.DOTALL)
    return fixed_text

def process_single_image(data):
    filename, img_base64 = data
    payload = {
        "model": MODEL_NAME,
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": [
                {"type": "text", "text": "Trích xuất JSON."},
                {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_base64}"}}
            ]}
        ],
        "stream": False,
        "max_tokens": 500,
        "top_p": 0.1
    }

    try:
        response = session.post(OLLAMA_URL, json=payload, timeout=60)
        response.raise_for_status()
        raw_text = response.json()["choices"][0]["message"]["content"]

        # 1. Fix lỗi dấu ngoặc kép trước
        cleaned_json_str = fix_json_quotes(raw_text)
        
        if cleaned_json_str:
            try:
                data_json = json.loads(cleaned_json_str)
                
                # 2. Xóa dấu ngoặc kép một lần nữa trong object cho chắc chắn
                for key in ["trichyeu", "domat"]:
                    if isinstance(data_json.get(key), str):
                        data_json[key] = data_json[key].replace('"', '')
                
                data_json["filename"] = filename
                result.append(data_json)
                
                # Log kết quả ĐÃ FIX để bạn xem
                log_console(filename, f"Cleaned JSON: {json.dumps(data_json, ensure_ascii=False)}")
                log_console(filename, "✅ Parse JSON thành công", "SUCCESS")
            except json.JSONDecodeError as e:
                log_console(filename, f"Raw AI failed: {raw_text}", "ERROR")
                log_console(filename, f"❌ Lỗi JSON sau khi fix: {e}", "ERROR")

    except Exception as e:
        log_console(filename, f"💥 Lỗi kết nối/API: {str(e)}", "CRITICAL")

# =========================
# 3️⃣ CHUẨN HÓA & EXCEL
# =========================
def result_process():
    for item in result:
        # Chuẩn hóa null
        for key, value in item.items():
            if value in ["null", "", "None", None]:
                item[key] = None
        
        # Chuẩn hóa ngày (giữ nguyên logic gốc của bạn)
        if item.get("ngay_banhanh"):
            item["ngay_banhanh"] = item["ngay_banhanh"].strip()
            if not re.match(r"\d{2}/\d{2}/\d{4}$", item["ngay_banhanh"]):
                item["ngay_banhanh"] = None

        # Chuẩn hóa độ mật (A, B, C)
        domat = item.get("domat")
        if not domat:
            item["domat"] = "Không"
        else:
            dm = domat.lower()
            if "tuyệt" in dm: item["domat"] = "A"
            elif "tối" in dm: item["domat"] = "B"
            elif "mật" in dm: item["domat"] = "C"
            else: item["domat"] = "Không"

def export_to_excel(data, output_file="ket_qua_ocr.xlsx"):
    full_path = os.path.join(BASE_DIR, output_file)
    wb = Workbook()
    ws = wb.active
    ws.append(["loai_vb", "domat", "so_vb", "ngay_banhanh", "trichyeu"])

    for item in data:
        ngay_dt = None
        if item.get("ngay_banhanh"):
            try:
                ngay_dt = datetime.strptime(item["ngay_banhanh"], "%d/%m/%Y")
            except: pass
        
        ws.append([item.get("loai_vb"), item.get("domat"), item.get("so_vb"), ngay_dt, item.get("trichyeu")])
        if ngay_dt:
            ws.cell(row=ws.max_row, column=4).number_format = "DD/MM/YYYY"

    wb.save(full_path)
    print(f"\n[!] Đã xuất kết quả ra file: {full_path}")

# =========================
# 4️⃣ MAIN
# =========================
def main():
    if not os.path.exists(IMAGE_FOLDER):
        print(f"Lỗi: Không tìm thấy thư mục {IMAGE_FOLDER}")
        return

    image_files = [f for f in os.listdir(IMAGE_FOLDER) if f.lower().endswith((".jpg", ".jpeg", ".png"))]
    if not image_files:
        print("Thư mục trống.")
        return

    print(f"--- Bắt đầu xử lý {len(image_files)} file ---")
    
    encoded_data = []
    for filename in image_files:
        img_64 = img_process(os.path.join(IMAGE_FOLDER, filename))
        if img_64: encoded_data.append((filename, img_64))

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = [executor.submit(process_single_image, d) for d in encoded_data]
        for future in as_completed(futures):
            future.result()

    result_process()
    export_to_excel(result)

if __name__ == "__main__":
    main()